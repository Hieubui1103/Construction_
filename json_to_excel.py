"""
"images_action/batch_1/image002.png": [
        "yes",
        "yes",
        "no",
        "no",
        "yes",
        "no",
        "yes",
        "no",
        "no",
        "no"
    ],

Given the following JSON data, convert it to an Excel file with the following specifications:
- Convert "yes" to 1 and "no" to 0
- if the answer is "unanswerable", use random 0 or 1 to fill in the cell
- Create an Excel sheet with 10 columns (indexes 0-9)
- The headers are given, just need to fill in the data from row 2
- There are 5 batches of images, each containing 200 images
- The images are named in the format "image001.jpg", "image002.jpg", ..., "image1000.jpg"
- Go over each batch and fill in the data accordingly
- The excel file is given as well as the placeholder sheet. Just need to fill them in. 

"""

import json
import random
import os
import pandas as pd
from openpyxl import load_workbook

# Create DataFrame with descriptive headers
headers = [ 
    [
    "Operating heavy machinery",
    "Lifting and carrying materials",
    "Measuring and marking surfaces",
    "Mixing cement or concrete",
    "Collaborating in teams",
    "Using hand tools (e.g., hammering, drilling)",
    "Inspecting completed work",
    "Communicating with supervisors",
    "Following safety protocols",
    "Taking breaks/resting",
    ] # Action labels
    ,
    [
    "Focused",
    "Determined",
    "Tired",
    "Alert",
    "Satisfied",
    "Anxious",
    "Proud",
    "Frustrated",
    "Cooperative",
    "Relieved"
    ] # Emotion labels
]

def json_to_excel(json_directory, headers, sheet_name, book_path, category):
   
    # Load JSON data from the directory
    json_data = {}

    def find_last_non_empty_row(ws):
        # Iterate over all rows and find the last row with content in any cell
        for row in range(ws.max_row, 0, -1):
            # Check if any cell in the row is non-empty
            if any(cell.value not in [None, ''] for cell in ws[row]):
                return row
        return 0

    for filename in os.listdir(json_directory):
        if filename.endswith(".json"):
            file_path = os.path.join(json_directory, filename)
            with open(file_path, "r") as json_file:
                data = json.load(json_file)
                json_data.update(data)

    rows = []
    for i in range(1, len(json_data) + 1):
        img_key = f"images_{category}/batch_{((i - 1) // 200) + 1}/image{str(i).zfill(3)}.png"
        if img_key in json_data:
            row = [1 if x == "yes" else 0 if x =="no" else random.randint(0, 1) for x in json_data[img_key]]
            rows.append(row)
        

    df = pd.DataFrame(rows,columns=headers)
    print(df)

    with pd.ExcelWriter(book_path, mode='a', engine='openpyxl',if_sheet_exists = 'overlay') as writer:
            wb = load_workbook(book_path)
            
            # Check if the sheet exists
            if sheet_name in wb.sheetnames:
                # If the sheet exists, get the last row with data
                ws = wb[sheet_name]
                #last_row = ws.max_row
                last_row = find_last_non_empty_row(ws)
                print(last_row)

            # Write the DataFrame starting at the next available row
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=last_row)
            
            # Load the workbook again after writing data
            wb = load_workbook(book_path)
            ws = wb[sheet_name]

            # Set column widths (auto-adjust based on header length)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = 30
                ws.column_dimensions[column].width = adjusted_width

            # Set row heights (example: automatic height for each row)
            for row in ws.iter_rows(min_row=last_row, max_row=last_row + len(rows)):
                for cell in row:
                    ws.row_dimensions[cell.row].height = 300  # Default height for all rows

            # Save the workbook with modifications
            wb.save(book_path)

if __name__ == "__main__":
    # Do each sheet of each excel file separately
    json_to_excel("vLLMs_ans/Florence_action", headers[0], "Florence_action", "Action_cons.xlsx", "action")
    json_to_excel("vLLMs_ans/Florence_emotion", headers[1], "Florence_emotion", "Emotion_cons.xlsx", "emotion")
    #json_to_excel("vLLMs_ans/LLaVa1.5_action", headers[0], "LLaVa1.5_action", "Action_cons.xlsx", "action")
    #json_to_excel("vLLMs_ans/LLaVa1.5_emotion", headers[1], "LLaVa1.5_emotion", "Emotion_cons.xlsx", "emotion")