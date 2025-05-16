import pandas as pd
import pickle
from openpyxl import load_workbook

def read_excel_to_matrix(sheet, file_path): 
    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name=sheet)
    
    # Extract only the first 10 columns
    df = df.iloc[:, :10]
    
    # Convert the DataFrame to a 2D matrix
    matrix = df.values.tolist()
    
    return matrix

def save_matrix_to_pickle(matrix, filename):
    with open(filename, 'wb') as f:
        pickle.dump(matrix, f)

def excel_to_pickle(file_path):
    # Load the workbook and get the sheet names
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames
    print(sheet_list)

    for sheet in sheet_list:
        matrix = read_excel_to_matrix(sheet, file_path)
        # Print the matrix and its length
        print(f"{sheet} matrix:")
        print(len(matrix))
    
         # Save the matrix to a pickle file
        save_matrix_to_pickle(matrix, f'../vLLMs_mat/{sheet}.pkl')

if __name__ == "__main__":
    file_path_action = '../Action_cons.xlsx' # Replace with the path to your Excel file
    file_path_emotion = '../Emotion_cons.xlsx' # Replace with the path to your Excel file
    excel_to_pickle(file_path_action)
    excel_to_pickle(file_path_emotion)
