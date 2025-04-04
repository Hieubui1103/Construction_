from openai import OpenAI
import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from aws import get_s3_file_urls
from dotenv import load_dotenv

load_dotenv()

client = OpenAI(
    # This is the default and can be omitted
    api_key= os.getenv("GPT_KEY")
)


# Import the Excel file
output_file = "action_label.xlsx"
        # Load the existing workbook
if os.path.exists(output_file):
    book = openpyxl.load_workbook(output_file)
else:
    book = openpyxl.Workbook()

    # Select the desired sheet (create a new one if needed)
if 'Responses' in book.sheetnames:
    sheet = book['Responses']
else:
    sheet = book.create_sheet('Responses')

new_header =  [
        "Operating heavy machinery",
        "Lifting and carrying materials",
        "Measuring and marking surfaces",
        "Mixing cement or concrete",
        "Collaborating in teams",
        "Using hand tools (e.g., hammering, drilling)",
        "Inspecting completed work",
        "Communicating with supervisors",
        "Following safety protocols",
        "Taking breaks/resting",
        "Photo link"
        ]

data = {header_item: None for header_item in new_header}

def get_last_non_empty_row(sheet):
    # Start from the last row and move upwards
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value is not None for cell in sheet[row]):
            return row
    return 0  # Return 0 if all rows are empty

def parse_gpt_response(response, img):
    """
    Parse the GPT-4 response to extract the 1s and 0s for each action.
    """
    # Initialize a list to store the parsed values
    parsed_response = []
    
    # Split the response into lines
    lines = response.split("\n")
    # Iterate through each line and extract the label (1 or 0)
    for line in lines:
        if "-" in line:  # Check if the line contains a label
            # Extract the part after the hyphen and strip whitespace
            label_part = line.split("-")[-1].strip()
            # Extract the first character (1 or 0) from the label part
            label = label_part.split()[0]  # Split by space and take the first part
            if label.isdigit(): 
                parsed_response.append(int(label))

    parsed_response.append(img)
    print(parsed_response)
    for i, header_item in enumerate(new_header):
        if data[header_item] is None:
            data[header_item] = [parsed_response[i]]  # Initialize with the first value as a list
        else:
            data[header_item].append(parsed_response[i])  # Append subsequent values as separate lists

def import_excel():
    global data
    global sheet
    global output_file
    global book
    
    last_row = get_last_non_empty_row(sheet)

    for i, header_item in enumerate(new_header):
        # Get the list of values for the current header
        values = data[header_item]

        # Write each value to the corresponding cell in the Excel sheet
        for j, value in enumerate(values):
            sheet.cell(row=last_row + 1 + j, column=i + 1).value = value  # Write the value directly
            sheet.cell(row=last_row + 1 + j, column=i + 1).alignment = Alignment(wrap_text=True)
            sheet.row_dimensions[last_row + 1 + j].height = 60
    
    book.save(output_file)

img_url = get_s3_file_urls()
print(len(img_url))
for i in range(1,len(img_url)):
    # if img_url == "q":
    #     break
    image_paths_input = img_url[i]
    #photo_dt.append(image_paths_input)
    print(f"\n For photo number {i}: ")
    print(image_paths_input)
    response = client.chat.completions.create(
        model="gpt-4-turbo",
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": """Identify the actions that appear in this photo, labeling as 1 and 0 for each of these actions: (use "-" to annotate 1 or 0 next to the label name: e.g. Operating heavy machinery - 1) (Just labelling them, no need further explanation for each action.) [
                    "Operating heavy machinery",
                    "Lifting and carrying materials",
                    "Measuring and marking surfaces",
                    "Mixing cement or concrete",
                    "Collaborating in teams",
                    "Using hand tools (e.g., hammering, drilling)",
                    "Inspecting completed work",
                    "Communicating with supervisors",
                    "Following safety protocols",
                    "Taking breaks/resting"
                    ]?"""},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"{image_paths_input}"},
                    },
                ],
            }
        ],
    )
    response_1 = response.choices[0].message.content
    print(response_1)
    parse_gpt = parse_gpt_response(response_1, image_paths_input)

import_excel()

