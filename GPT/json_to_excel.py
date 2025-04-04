import json
import pandas as pd
from openpyxl import load_workbook


def find_last_non_empty_row(ws):
    # Iterate over all rows and find the last row with content in any cell
    for row in range(ws.max_row, 0, -1):
        # Check if any cell in the row is non-empty
        if any(cell.value not in [None, ''] for cell in ws[row]):
            return row
    return 0

def json_to_excel(json_path, sheet_name, book_path = '../action_label.xlsx'):
    """
    Convert JSON data to Excel format with specific requirements:
    - Convert "yes" to 1 and "no" to 0
    - Create Excel sheet with 10 columns (indexes 0-9)
    - Add headers for columns 0-9
    - Save to action_label.xlsx in "LLaVa-1.5-action" sheet
    """
    # Strip quotes from path if present
    # json_path = json_path.strip('"\'')
    
    # Load JSON data
    with open(json_path, 'r') as f:
        data = json.load(f)

    # Create DataFrame with converted values
    rows = []
    for i in range(1, 105):
        img_key = f"images/image{str(i).zfill(3)}.jpg"
        if img_key in data:
            row = [1 if x == "Yes" else 0 for x in data[img_key]]
            rows.append(row)
    
    # Create DataFrame with descriptive headers
    headers = [
        "Operating heavy machinery",
        "Lifting and carrying materials",
        "Measuring and marking surfaces",
        "Mixing cement or concrete",
        "Collaborating in teams",
        "Using hand tools (e.g., hammering, drilling)",
        "Inspecting completed work",
        "Communicating with supervisors",
        "Following safety protocols",
        "Taking breaks/resting",
        ]
    
    # headers = [
    #     "Focused",
    #     "Determined",
    #     "Tired",
    #     "Alert",
    #     "Satisfied",
    #     "Anxious",
    #     "Proud",
    #     "Frustrated",
    #     "Cooperative",
    #     "Relieved"
    #     ]

    df = pd.DataFrame(rows,columns=headers)
    print(df)
    
    # Save to Excel
    with pd.ExcelWriter(book_path, mode='a', engine='openpyxl',if_sheet_exists = 'overlay') as writer:
        wb = load_workbook(book_path)
        
        # Check if the sheet exists
        if sheet_name in wb.sheetnames:
            # If the sheet exists, get the last row with data
            ws = wb[sheet_name]
            #last_row = ws.max_row
            last_row = find_last_non_empty_row(ws)
            print(last_row)

        # Write the DataFrame starting at the next available row
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=last_row)
        
        # Load the workbook again after writing data
        wb = load_workbook(book_path)
        ws = wb[sheet_name]

        # Set column widths (auto-adjust based on header length)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = 30
            ws.column_dimensions[column].width = adjusted_width

        # Set row heights (example: automatic height for each row)
        for row in ws.iter_rows(min_row=last_row, max_row=last_row + len(rows)):
            for cell in row:
                ws.row_dimensions[cell.row].height = 300  # Default height for all rows

        # Save the workbook with modifications
        wb.save(book_path)

if __name__ == "__main__":
    json_path = r"..\vLLMs_ans\response_action_LLaVa_1.json" # Replace with your own path 
    json_to_excel(json_path, sheet_name="LLaVa-1.5-actions")
