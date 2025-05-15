import base64
from openai import OpenAI
from PIL import Image
import os
import openpyxl
from openpyxl.styles import Alignment


from dotenv import load_dotenv

load_dotenv()

client = OpenAI(
    # This is the default and can be omitted
    api_key=os.getenv('GPT_KEY'),
)


# Import the Excel file
output_file = "Emotion_cons.xlsx"
        # Load the existing workbook
if os.path.exists(output_file):
    book = openpyxl.load_workbook(output_file)
else:
    book = openpyxl.Workbook()

    # Select the desired sheet (create a new one if needed)
if 'GPT_Emotions' in book.sheetnames:
    sheet = book['GPT_Emotions']
else:
    sheet = book.create_sheet('GPT_Emotions')

new_header =  [
        "Focused",
        "Determined",
        "Tired",
        "Alert",
        "Satisfied",
        "Anxious",
        "Proud",
        "Frustrated",
        "Cooperative",
        "Relieved",
        ]

data = {header_item: None for header_item in new_header}

def get_last_non_empty_row(sheet):
    # Start from the last row and move upwards
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value is not None for cell in sheet[row]):
            return row
    return 0  # Return 0 if all rows are empty

def parse_gpt_response(response):
    """
    Parse the GPT-4 response to extract the 1s and 0s for each action.
    """
    # Initialize a list to store the parsed values
    parsed_response = []
    
    # Split the response into lines
    lines = response.split("\n")
    # Iterate through each line and extract the label (1 or 0)
    for line in lines:
        if "-" in line:  # Check if the line contains a label
            # Extract the part after the hyphen and strip whitespace
            label_part = line.split("-")[-1].strip()
            # Extract the first character (1 or 0) from the label part
            label = label_part.split()[0]  # Split by space and take the first part
            if label.isdigit(): 
                parsed_response.append(int(label))

    print(parsed_response)
    for i, header_item in enumerate(new_header):
        if data[header_item] is None:
            data[header_item] = [parsed_response[i]]  # Initialize with the first value as a list
        else:
            data[header_item].append(parsed_response[i])  # Append subsequent values as separate lists

def import_excel():
    global data
    global sheet
    global output_file
    global book
    
    last_row = get_last_non_empty_row(sheet)

    for i, header_item in enumerate(new_header):
        # Get the list of values for the current header
        values = data[header_item]

        # Write each value to the corresponding cell in the Excel sheet
        for j, value in enumerate(values):
            sheet.cell(row=last_row + 1 + j, column=i + 1).value = value  # Write the value directly
            sheet.cell(row=last_row + 1 + j, column=i + 1).alignment = Alignment(wrap_text=True)
            sheet.row_dimensions[last_row + 1 + j].height = 60
    
    book.save(output_file)
# Function to encode the image
def encode_image(image_path):
    # Encode as base64
    with open(image_path, "rb") as image_file:
        image_data = image_file.read()
        if image_data[:8] != b"\x89PNG\r\n\x1a\n":
            raise ValueError("File is not a valid PNG")
        return base64.b64encode(image_data).decode("utf-8")


# Path to your image
for folder in os.listdir(r"C:\VScode\Construction_CV\photo_extract\emotion_images"):
    if folder.endswith(".zip"):
        continue
    path_batch = os.path.join(r"C:\VScode\Construction_CV\photo_extract\emotion_images",folder)
    for file in os.listdir(path_batch):
        if file.endswith(".png"):
            image_path = os.path.join(path_batch, file)
            
        print(f"\n For photo number {file}: ")
        # Getting the Base64 string
        base64_image = encode_image(image_path)

        response = client.responses.create(
            model="chatgpt-4o-latest",
            input=[
                {
                    "role": "user",
                    "content": [
                        { "type": "input_text", 
                        "text":  """Identify the emotions that are detected in this photo, labeling as 1 or 0 for each of these identified emotions: (use "-" to annotate 1 or 0 next to the label name: e.g. Focused - 1) If there arent any human figures or faces or you cannot identify, just label "0" for all of them. (Also, just labelling them, no need further explanation for each emotion.) [
                        "Focused",
                        "Determined",
                        "Tired",
                        "Alert",
                        "Satisfied",
                        "Anxious",
                        "Proud",
                        "Frustrated",
                        "Cooperative",
                        "Relieved",
                        ]?"""},
                        {
                            "type": "input_image",
                            "image_url": f"data:image/png;base64,{base64_image}",
                        },
                    ],
                }
            ],
        )
        response_1 = response.output_text
        print(response.output_text)
        parse_gpt = parse_gpt_response(response_1)

import_excel()

# Identify the actions that appear in this photo, labeling as 1 and 0 for each of these actions: (use "-" to annotate 1 or 0 next to the label name: e.g. Operating heavy machinery - 1) (Just labelling them, no need further explanation for each action.). If there is no match, labelling all the actions with 0s'.[
#                         "Operating heavy machinery",
#                         "Lifting and carrying materials",
#                         "Measuring and marking surfaces",
#                         "Mixing cement or concrete",
#                         "Collaborating in teams",
#                         "Using hand tools (e.g., hammering, drilling)",
#                         "Inspecting completed work",
#                         "Communicating with supervisors",
#                         "Following safety protocols",
#                         "Taking breaks/resting"
#                         ]?



